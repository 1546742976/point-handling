#!/usr/bin/env python3
"""
丝绸之路轨迹数据生成器
========================
使用自然三次样条插值（C² 连续）生成等弧长采样的路径轨迹。
通过系数混合法（自然样条 ↔ 折线）调节弧长，轨迹始终经过所有控制点。
输出 N 行制表符分隔的 (X, Y, 有符号曲率半径 R) 数据。

用法:
    python generate_trajectory.py [--plot] [--smooth S] [--n-output N] [--target-length L] <输入文件> [输出文件]
    python generate_trajectory.py --plot --smooth 0.3                     # stdin + 平滑 + 可视化
    python generate_trajectory.py                                         # stdin（仅数据）

选项:
    --plot              生成 PNG 轨迹地图可视化
    --xlsx              同时导出 Excel (.xlsx) 文件，包含轨迹数据、控制点和参数
    --smooth S          控制点平滑强度，范围 [0, 1]，默认 0（不平滑）
    --n-output N        等弧长采样点数，默认 2000
    --target-length L   期望轨迹总长度，> 0。通过系数混合（二分搜索 α）实现，
                        轨迹始终经过所有控制点；弦长总和为理论最小弧长

输入文件格式:
    第一行: 地图宽度 地图高度
    第二行: 控制点数量 N
    后续 N 行: 控制点坐标 X Y（空格/逗号/制表符分隔）
    空行和 # 开头的行为注释

输出:
    N 行纯文本，无表头，制表符分隔，每行: X  Y  R
    所有数值保留两位小数
"""

import sys
import numpy as np


# ══════════════════════════════════════════════════
# 1. 输入：读取控制点
# ══════════════════════════════════════════════════

def _parse_line(line):
    """解析一行文本，返回浮点数列表。空行/注释返回 None。"""
    line = line.strip()
    if not line or line.startswith('#'):
        return None
    parts = line.replace(',', ' ').replace('\t', ' ').split()
    return [float(x) for x in parts] if parts else None


def read_map_and_points(filepath=None):
    """
    从文件或 stdin 读取地图尺寸、控制点数量与控制点。

    文件格式:
        第 1 个有效数据行: 地图宽度 地图高度
        第 2 个有效数据行: 控制点数量 N
        后续 N 行: 控制点 X Y

    stdin 模式: 若实际读取点数与声明不一致，提示重新输入数量+坐标。
    """
    is_stdin = (filepath is None)

    # ── 读取地图尺寸 ──
    map_size = None
    if is_stdin:
        while True:
            line = sys.stdin.readline()
            if line == '':
                break
            vals = _parse_line(line)
            if vals is None:
                continue
            if len(vals) >= 2:
                map_size = (vals[0], vals[1])
                break
    else:
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                vals = _parse_line(line)
                if vals is None:
                    continue
                if len(vals) >= 2:
                    map_size = (vals[0], vals[1])
                    break

    if map_size is None:
        print("错误：第一行必须输入地图尺寸（宽度 高度）", file=sys.stderr)
        sys.exit(1)

    # ── 文件模式：一次性读取剩余数据 ──
    if not is_stdin:
        n_expected = None
        points = []
        with open(filepath, 'r', encoding='utf-8') as f:
            header_done = 0  # 0=等地图, 1=等数量, 2=等坐标
            for line in f:
                vals = _parse_line(line)
                if vals is None:
                    continue
                if header_done == 0:
                    header_done = 1
                    continue
                if header_done == 1:
                    if len(vals) >= 1:
                        n_expected = int(vals[0])
                        header_done = 2
                    continue
                if len(vals) >= 2:
                    points.append((vals[0], vals[1]))

        if n_expected is None:
            print("错误：第二行必须输入控制点数量", file=sys.stderr)
            sys.exit(1)
        if len(points) < 2:
            print("错误：至少需要 2 个控制点", file=sys.stderr)
            sys.exit(1)
        if len(points) != n_expected:
            print(f"警告：声明的控制点数量 ({n_expected}) 与实际读取数量 ({len(points)}) 不一致",
                  file=sys.stderr)
        return map_size, np.array(points, dtype=np.float64)

    # ── stdin 模式：循环读取（不一致时允许重新输入） ──
    while True:
        # 读取控制点数量
        print("请输入控制点数量 N:", file=sys.stderr)
        n_line = sys.stdin.readline()
        if n_line == '':  # EOF
            print("错误：未输入控制点数量", file=sys.stderr)
            sys.exit(1)
        while True:
            vals = _parse_line(n_line)
            if vals is not None and len(vals) >= 1:
                n_expected = int(vals[0])
                break
            print("请输入控制点数量 N:", file=sys.stderr)
            n_line = sys.stdin.readline()
            if n_line == '':  # EOF
                print("错误：未输入控制点数量", file=sys.stderr)
                sys.exit(1)

        # 读取 N 个控制点
        print(f"请输入 {n_expected} 个控制点坐标（每行 X Y）:", file=sys.stderr)
        points = []
        for i in range(n_expected):
            line = sys.stdin.readline()
            if line == '':  # 提前 EOF
                break
            vals = _parse_line(line)
            while vals is None or len(vals) < 2:
                print(f"  第 {i+1}/{n_expected} 个点格式有误，请重新输入（X Y）:", file=sys.stderr)
                line = sys.stdin.readline()
                if line == '':
                    break
                vals = _parse_line(line)
            if line == '':
                break
            points.append((vals[0], vals[1]))

        if len(points) < 2:
            print("错误：至少需要 2 个控制点", file=sys.stderr)
            sys.exit(1)

        if len(points) == n_expected:
            break

        print(f"警告：声明的控制点数量 ({n_expected}) 与实际读取数量 ({len(points)}) 不一致，"
              f"请重新输入。", file=sys.stderr)

    return map_size, np.array(points, dtype=np.float64)


# ══════════════════════════════════════════════════
# 2. 控制点平滑预处理
# ══════════════════════════════════════════════════

def smooth_control_points(pts, strength=0.0):
    """
    使用拉普拉斯平滑（Laplacian smoothing）预处理控制点。

    每个内部控制点向其相邻点中点的方向移动，端点不动。
    适用于减少三次样条在尖锐转向处的过冲与尖角。

    参数:
        pts      : shape (n, 2)  原始控制点
        strength : float, [0, 1]  平滑强度
                   0 = 不平滑（保持原样）
                   1 = 完全移到相邻点中点

    返回:
        shape (n, 2)  平滑后的控制点
    """
    if strength <= 0.0 or len(pts) <= 2:
        return pts.copy()

    result = pts.astype(np.float64).copy()
    n = len(result)
    for i in range(1, n - 1):
        midpoint = (result[i - 1] + result[i + 1]) / 2.0
        result[i] = result[i] + strength * (midpoint - result[i])

    return result


def smooth_output_trajectory(x_out, y_out, ctrl_pts, strength, n_iter=5):
    """
    对输出轨迹点做 Laplacian 平滑，同时保持控制点位置不变。

    在等弧长输出点（~2000 点）上施加平滑，每轮迭代后强制控制点位置还原。
    平滑完成后重新等弧长分配，保证输出点间距均匀。

    与 smooth_control_points（预平滑控制点）不同，此函数保证无论平滑
    强度多大，最终轨迹始终经过所有原始控制点。

    参数:
        x_out    : shape (n_out,)  等弧长输出轨迹 X 坐标
        y_out    : shape (n_out,)  等弧长输出轨迹 Y 坐标
        ctrl_pts : shape (n_ctrl, 2)  原始控制点
        strength : float, [0, 1]  平滑强度
                   0 = 不处理
                   1 = 最大平滑（多次迭代后趋近折线）
        n_iter   : int  平滑迭代次数

    返回:
        (x_sm, y_sm)  平滑 + 重新等弧长后的输出轨迹点
    """
    if strength <= 0.0 or len(ctrl_pts) <= 2:
        return x_out.copy(), y_out.copy()

    n_out = len(x_out)
    ctrl_arr = np.asarray(ctrl_pts, dtype=np.float64)

    # ── 找到每个控制点对应的最近输出点索引 ──
    fixed = {}  # idx → (cx, cy)
    for cp in ctrl_arr:
        dist2 = (x_out - cp[0]) ** 2 + (y_out - cp[1]) ** 2
        idx = int(np.argmin(dist2))
        fixed[idx] = (float(cp[0]), float(cp[1]))

    x_res = x_out.astype(np.float64).copy()
    y_res = y_out.astype(np.float64).copy()

    # ── 迭代平滑 ──
    for _ in range(n_iter):
        x_new = x_res.copy()
        y_new = y_res.copy()

        # 向量化 Laplacian：内部点移向相邻点中点
        x_new[1:-1] = x_res[1:-1] + strength * (
            (x_res[:-2] + x_res[2:]) * 0.5 - x_res[1:-1]
        )
        y_new[1:-1] = y_res[1:-1] + strength * (
            (y_res[:-2] + y_res[2:]) * 0.5 - y_res[1:-1]
        )

        # 强制控制点位置精确还原
        for idx, (cx, cy) in fixed.items():
            x_new[idx] = cx
            y_new[idx] = cy

        x_res, y_res = x_new, y_new

    # ── 重新等弧长分配 ──
    dx = np.diff(x_res)
    dy = np.diff(y_res)
    ds = np.sqrt(dx ** 2 + dy ** 2)
    arc = np.zeros(n_out, dtype=np.float64)
    arc[1:] = np.cumsum(ds)
    total_len = arc[-1]

    target_arc = np.linspace(0.0, total_len, n_out)
    x_eq = np.interp(target_arc, arc, x_res)
    y_eq = np.interp(target_arc, arc, y_res)

    # 重新等弧长后控制点可能有微小偏移，再次修正
    for cp in ctrl_arr:
        dist2 = (x_eq - cp[0]) ** 2 + (y_eq - cp[1]) ** 2
        idx = int(np.argmin(dist2))
        x_eq[idx] = cp[0]
        y_eq[idx] = cp[1]

    return x_eq, y_eq


# ══════════════════════════════════════════════════
# 3. 三次样条拟合（自然边界条件）
# ══════════════════════════════════════════════════

def fit_natural_cubic_spline(t_knots, y):
    """
    用自然边界条件（两端二阶导数为 0）拟合三次样条。

    参数:
        t_knots : shape (n+1,)  参数节点（累积弦长）
        y       : shape (n+1,)  待插值的函数值

    返回:
        a, b, c, d : shape (n,)  每段多项式的系数

    区间 i（对应 [t_i, t_{i+1}]）上:
        S_i(t) = a_i + b_i·(t-t_i) + c_i·(t-t_i)² + d_i·(t-t_i)³
    """
    n_seg = len(t_knots) - 1
    h = np.diff(t_knots)

    if n_seg == 1:
        # 只有两个点：退化为线性
        a = np.array([y[0]])
        b = np.array([(y[1] - y[0]) / h[0]])
        c = np.array([0.0])
        d = np.array([0.0])
        return a, b, c, d

    # ── 构建三对角系统求解 M_i = S''(t_i) ──
    # i = 1, …, n_seg-1（内点）:
    #   μ_i·M_{i-1} + 2·M_i + λ_i·M_{i+1} = d_i
    #   μ_i = h_{i-1}/(h_{i-1}+h_i),  λ_i = h_i/(h_{i-1}+h_i)
    #   d_i = 6/(h_{i-1}+h_i) · [(y_{i+1}-y_i)/h_i - (y_i-y_{i-1})/h_{i-1}]
    # 自然边界: M_0 = M_n = 0

    n_eq = n_seg - 1
    diag = np.ones(n_eq) * 2.0
    sub_diag = np.zeros(n_eq - 1)
    sup_diag = np.zeros(n_eq - 1)
    rhs = np.zeros(n_eq)

    for i in range(1, n_seg):
        row = i - 1
        h_left = h[i - 1]
        h_right = h[i]
        denom = h_left + h_right

        if row > 0:
            sub_diag[row - 1] = h_left / denom
        if row < n_eq - 1:
            sup_diag[row] = h_right / denom

        rhs[row] = (6.0 / denom) * (
            (y[i + 1] - y[i]) / h_right - (y[i] - y[i - 1]) / h_left
        )

    # 用 Thomas 算法解三对角系统（比 np.linalg.solve 更稳定）
    M_inner = _tridiagonal_solve(sub_diag.copy(), diag.copy(), sup_diag.copy(), rhs.copy())

    M = np.zeros(n_seg + 1)
    M[1:-1] = M_inner

    # ── 计算每段的多项式系数 ──
    a = y[:-1].copy()
    b = np.zeros(n_seg)
    c = np.zeros(n_seg)
    d = np.zeros(n_seg)

    for i in range(n_seg):
        hi = h[i]
        b[i] = (y[i + 1] - y[i]) / hi - hi * (2.0 * M[i] + M[i + 1]) / 6.0
        c[i] = M[i] / 2.0
        d[i] = (M[i + 1] - M[i]) / (6.0 * hi)

    return a, b, c, d


def _tridiagonal_solve(sub_diag, diag, sup_diag, rhs):
    """
    Thomas 算法求解三对角线性系统。

    sub_diag: 下对角线，长度 n-1
    diag:     主对角线，长度 n
    sup_diag: 上对角线，长度 n-1
    rhs:      右端项，长度 n
    """
    n = len(diag)
    for i in range(1, n):
        w = sub_diag[i - 1] / diag[i - 1]
        diag[i] -= w * sup_diag[i - 1]
        rhs[i] -= w * rhs[i - 1]

    x = np.zeros(n)
    x[-1] = rhs[-1] / diag[-1]
    for i in range(n - 2, -1, -1):
        x[i] = (rhs[i] - sup_diag[i] * x[i + 1]) / diag[i]

    return x


# ══════════════════════════════════════════════════
# 3. 样条求值（逐段掩码，稳健无 fancy-indexing 风险）
# ══════════════════════════════════════════════════

def eval_spline(t_vals, t_knots, a, b, c, d, deriv=0):
    """
    在参数值处评估样条或其导数。使用逐段布尔掩码，回避
    searchsorted + fancy indexing 在某些 numpy 版本 / 输入
    形状下的不确定行为。

    参数:
        t_vals  : 标量或数组，参数值
        t_knots : 参数节点
        a,b,c,d : 样条系数
        deriv   : 0=函数值, 1=一阶导, 2=二阶导

    返回:
        与输入同 shape 的评估结果
    """
    t = np.atleast_1d(np.asarray(t_vals, dtype=np.float64))
    scalar_input = (np.asarray(t_vals).ndim == 0)
    result = np.zeros(len(t), dtype=np.float64)
    n_seg = len(a)

    for seg in range(n_seg):
        lo = t_knots[seg]
        hi = t_knots[seg + 1]

        # 区间 [lo, hi)，最后一个区间包含右端点
        mask = (t >= lo) & (t < hi)
        if seg == n_seg - 1:
            mask = mask | (np.abs(t - hi) < 1e-14)

        if not np.any(mask):
            continue

        dt = t[mask] - lo

        if deriv == 0:
            result[mask] = a[seg] + b[seg] * dt + c[seg] * dt ** 2 + d[seg] * dt ** 3
        elif deriv == 1:
            result[mask] = b[seg] + 2.0 * c[seg] * dt + 3.0 * d[seg] * dt ** 2
        elif deriv == 2:
            result[mask] = 2.0 * c[seg] + 6.0 * d[seg] * dt
        else:
            raise ValueError("deriv 必须为 0, 1, 或 2")

    return float(result[0]) if scalar_input else result


# ══════════════════════════════════════════════════
# 3.6 轻量弧长计算（用于二分搜索，不需完整重采样）
# ══════════════════════════════════════════════════

def compute_total_arc_length(t_knots, ax, bx, cx, dx, ay, by, cy, dy,
                             n_dense=20000):
    """
    计算样条曲线的总弧长（仅密集采样 + 积分，不做等弧长插值）。

    参数与 arc_length_reparam 前四项相同。
    返回: float 总弧长
    """
    t_dense = np.linspace(t_knots[0], t_knots[-1], n_dense)
    x_dense = eval_spline(t_dense, t_knots, ax, bx, cx, dx, deriv=0)
    y_dense = eval_spline(t_dense, t_knots, ay, by, cy, dy, deriv=0)
    ds = np.sqrt(np.diff(x_dense) ** 2 + np.diff(y_dense) ** 2)
    return float(np.sum(ds))


# ══════════════════════════════════════════════════
# 3.7 样条系数混合：在自然样条与折线之间调节弧长
# ══════════════════════════════════════════════════

def blend_spline_coeffs(t_knots, ax, bx, cx, dx, y_ctrl, alpha):
    """
    将自然三次样条系数与折线系数按 α 混合。

    折线系数: a = y_i,  b_poly = (y_{i+1}-y_i)/h_i,  c = 0,  d = 0
    自然样条: a = y_i,  b, c, d 由 fit_natural_cubic_spline 给出

    混合规则:
        a_blend = a_spline (不变 → 始终经过所有控制点)
        b_blend = b_spline + α * (b_polyline - b_spline)
        c_blend = (1 - α) * c_spline
        d_blend = (1 - α) * d_spline

    α = 0:  纯自然样条（C² 连续，默认弧长）
    α > 0:  趋向折线（更直更短），α=1 → 折线（最短，C⁰）
    α < 0:  比自然样条更弯（更长）

    返回:
        ax_blend, bx_blend, cx_blend, dx_blend
    """
    n_seg = len(ax)
    h = np.diff(t_knots)

    ax_b = ax.copy()  # a 不变
    bx_b = np.zeros(n_seg, dtype=np.float64)
    cx_b = np.zeros(n_seg, dtype=np.float64)
    dx_b = np.zeros(n_seg, dtype=np.float64)

    for i in range(n_seg):
        hi = h[i]
        b_poly = (y_ctrl[i + 1] - y_ctrl[i]) / hi if hi > 1e-15 else 0.0

        bx_b[i] = bx[i] + alpha * (b_poly - bx[i])
        cx_b[i] = (1.0 - alpha) * cx[i]
        dx_b[i] = (1.0 - alpha) * dx[i]

    return ax_b, bx_b, cx_b, dx_b


def find_alpha_for_target_length(t_knots,
                                 ax_s, bx_s, cx_s, dx_s,
                                 ay_s, by_s, cy_s, dy_s,
                                 x_ctrl, y_ctrl,
                                 target_length,
                                 chord_sum,
                                 n_dense=20000,
                                 alpha_lo=-2.0, alpha_hi=1.0,
                                 max_iter=60, tol_rel=2e-4, tol_abs=0.5):
    """
    二分搜索 α 使混合样条总弧长 ≈ target_length。

    L(α) 随 α 增大而单调递减：
        α = 0  → 自然样条弧长
        α > 0  → 更直更短，α=1 → 折线（弦长总和）
        α < 0  → 更弯更长

    参数:
        alpha_lo : 二分下界（产生较长弧长）
        alpha_hi : 二分上界（产生较短弧长）
        chord_sum: 弦长总和（理论最小弧长）

    返回:
        (alpha, actual_length, success_flag, message)
    """
    if target_length <= chord_sum + 1e-6:
        msg = (f"⚠ 目标长度 ({target_length:.2f}) ≤ 弦长总和 ({chord_sum:.2f})，"
               f"折线为最短可能路径。使用弦长总和。")
        return 1.0, chord_sum, False, msg

    def arc_len_for_alpha(a):
        """给定 α，返回混合样条总弧长。"""
        ax_b, bx_b, cx_b, dx_b = blend_spline_coeffs(
            t_knots, ax_s, bx_s, cx_s, dx_s, x_ctrl, a)
        ay_b, by_b, cy_b, dy_b = blend_spline_coeffs(
            t_knots, ay_s, by_s, cy_s, dy_s, y_ctrl, a)
        return compute_total_arc_length(t_knots, ax_b, bx_b, cx_b, dx_b,
                                        ay_b, by_b, cy_b, dy_b, n_dense)

    # 检查 α_hi 处的弧长（应 < target）
    L_hi = arc_len_for_alpha(alpha_hi)
    if L_hi > target_length:
        msg = (f"⚠ 折线弧长 ({chord_sum:.2f}) 为最短可能路径，"
               f"无法达到目标 ({target_length:.2f})。使用折线。")
        return 1.0, chord_sum, False, msg

    # 检查 α_lo 处的弧长（应 > target）
    L_lo = arc_len_for_alpha(alpha_lo)
    expand_attempts = 0
    while L_lo < target_length and expand_attempts < 8:
        alpha_lo -= 2.0
        L_lo = arc_len_for_alpha(alpha_lo)
        expand_attempts += 1

    if L_lo < target_length:
        msg = (f"⚠ 即使在 α={alpha_lo:.1f} 处弧长 ({L_lo:.2f}) 仍小于目标 "
               f"({target_length:.2f})。使用最大可达弧长。")
        return alpha_lo, L_lo, False, msg

    # 二分搜索
    lo, hi = alpha_lo, alpha_hi

    for iteration in range(max_iter):
        alpha_mid = (lo + hi) / 2.0
        L_mid = arc_len_for_alpha(alpha_mid)

        error = abs(L_mid - target_length)
        tol = max(tol_rel * target_length, tol_abs)
        # 至少迭代 3 次再允许早停，避免初值巧合误判
        if error <= tol and iteration >= 3:
            return alpha_mid, L_mid, True, ""

        if L_mid < target_length:
            hi = alpha_mid
        else:
            lo = alpha_mid

    alpha_final = (lo + hi) / 2.0
    L_final = arc_len_for_alpha(alpha_final)
    err = abs(L_final - target_length)
    if err <= max(tol_rel * target_length, tol_abs):
        return alpha_final, L_final, True, ""
    msg = (f"⚠ 二分搜索达到最大迭代次数 ({max_iter})，"
           f"当前误差: {err:.2f}")
    return alpha_final, L_final, False, msg


# ══════════════════════════════════════════════════
# 4. 等弧长重参数化
# ══════════════════════════════════════════════════

def arc_length_reparam(t_knots, ax, bx, cx, dx, ay, by, cy, dy,
                       n_output=2000, n_dense=20000):
    """
    对样条曲线做等弧长重采样。

    策略（两步法，避免在弧长插值得到的 t 上评估样条）:
      1. 在密集的 t 网格上计算 x, y, x', y', x'', y''
      2. 计算累积弧长
      3. 用 np.interp 直接在弧长上插值所有量 → 等弧长采样点

    返回:
        x_out, y_out, xp, yp, xpp, ypp
    """
    # ── 密集参数采样 ──
    t_dense = np.linspace(t_knots[0], t_knots[-1], n_dense)

    x_dense = eval_spline(t_dense, t_knots, ax, bx, cx, dx, deriv=0)
    y_dense = eval_spline(t_dense, t_knots, ay, by, cy, dy, deriv=0)
    xp_dense = eval_spline(t_dense, t_knots, ax, bx, cx, dx, deriv=1)
    yp_dense = eval_spline(t_dense, t_knots, ay, by, cy, dy, deriv=1)
    xpp_dense = eval_spline(t_dense, t_knots, ax, bx, cx, dx, deriv=2)
    ypp_dense = eval_spline(t_dense, t_knots, ay, by, cy, dy, deriv=2)

    # ── 累积弧长 ──
    dx = np.diff(x_dense)
    dy = np.diff(y_dense)
    ds = np.sqrt(dx ** 2 + dy ** 2)
    arc_dense = np.zeros(n_dense, dtype=np.float64)
    arc_dense[1:] = np.cumsum(ds)
    total_len = arc_dense[-1]

    # ── 等弧长目标 ──
    target_arc = np.linspace(0.0, total_len, n_output)

    # ── 直接在弧长上线性插值所有量 ──
    # (20000 点密集采样 → 线性插值精度足够)
    x_out = np.interp(target_arc, arc_dense, x_dense)
    y_out = np.interp(target_arc, arc_dense, y_dense)
    xp = np.interp(target_arc, arc_dense, xp_dense)
    yp = np.interp(target_arc, arc_dense, yp_dense)
    xpp = np.interp(target_arc, arc_dense, xpp_dense)
    ypp = np.interp(target_arc, arc_dense, ypp_dense)

    return x_out, y_out, xp, yp, xpp, ypp


# ══════════════════════════════════════════════════
# 5. 有符号曲率半径
# ══════════════════════════════════════════════════

def compute_signed_radius(xp, yp, xpp, ypp):
    """
    计算有符号曲率半径。

    曲率 κ = (x'·y'' - y'·x'') / (x'² + y'²)^(3/2)
    （导数对任意参数 t，公式不变）

    半径 R = 1 / κ

    符号约定: 左转（逆时针）→ κ > 0 → R > 0
              右转（顺时针）→ κ < 0 → R < 0

    |κ| < 1e-8 → 直线段: R = ±99999.99
    """
    n = len(xp)
    numerator = xp * ypp - yp * xpp
    denominator = (xp ** 2 + yp ** 2) ** 1.5

    kappa = np.zeros(n, dtype=np.float64)
    R = np.zeros(n, dtype=np.float64)

    for i in range(n):
        denom_i = denominator[i]
        if abs(denom_i) < 1e-15:
            kappa[i] = 0.0
        else:
            kappa[i] = numerator[i] / denom_i

        if abs(kappa[i]) < 1e-8:
            sign = _find_nearest_sign(kappa, i)
            R[i] = sign * 99999.99
        else:
            R[i] = 1.0 / kappa[i]

    return R


def _find_nearest_sign(kappa, idx):
    """在 kappa 数组中查找距离 idx 最近的非零曲率的符号。"""
    n = len(kappa)
    for d in range(1, n):
        fwd = idx + d
        if fwd < n and abs(kappa[fwd]) >= 1e-8:
            return 1.0 if kappa[fwd] > 0 else -1.0
        bwd = idx - d
        if bwd >= 0 and abs(kappa[bwd]) >= 1e-8:
            return 1.0 if kappa[bwd] > 0 else -1.0
    return 1.0  # 所有点都是直线（极端情况）


# ══════════════════════════════════════════════════
# 6. 输出
# ══════════════════════════════════════════════════

def output_results(x, y, R, filepath=None):
    """输出 2000 行制表符分隔数据，保留两位小数。"""
    out = open(filepath, 'w', encoding='utf-8') if filepath else sys.stdout
    with out as f:
        for i in range(len(x)):
            f.write(f"{x[i]:.2f}\t{y[i]:.2f}\t{R[i]:.2f}\n")


def output_xlsx(filepath, x_traj, y_traj, R_traj, ctrl_pts, params=None):
    """
    将轨迹数据导出为 Excel (.xlsx) 文件。

    输出行数与 n_output 参数一致（全部等弧长采样点）。

    包含:
        - 轨迹数据: 序号, X, Y, 曲率半径 R
        - 生成参数: 参数名, 参数值

    参数:
        filepath  : xlsx 输出路径
        x_traj, y_traj, R_traj : 轨迹坐标和曲率半径
        ctrl_pts  : shape (n, 2)  控制点坐标
        params    : dict  生成参数信息
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 样式定义 ──
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2A78D6', end_color='2A78D6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    data_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    header_border = Border(
        left=Side(style='thin', color='1A5CB0'),
        right=Side(style='thin', color='1A5CB0'),
        top=Side(style='thin', color='1A5CB0'),
        bottom=Side(style='thin', color='1A5CB0'),
    )

    # ════════════════════════════════════════════════
    # Sheet 1: 轨迹数据（全部等弧长采样点）
    # ════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = '轨迹数据'

    n = len(x_traj)

    headers1 = ['序号', 'X', 'Y', '曲率半径 R']
    cols1 = [list(range(1, n + 1)), x_traj, y_traj, R_traj]

    # 写表头
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # 写数据
    for row_idx in range(n):
        for col_idx in range(len(headers1)):
            val = cols1[col_idx][row_idx]
            cell = ws1.cell(row=row_idx + 2, column=col_idx + 1,
                            value=round(float(val), 6) if abs(float(val)) < 1e6 else float(val))
            cell.alignment = data_align
            cell.border = thin_border
            cell.font = Font(name='Consolas', size=10)

    # 列宽
    col_widths1 = [6, 10, 10, 14]
    for col_idx, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = w

    # 冻结表头
    ws1.freeze_panes = 'A2'

    # ════════════════════════════════════════════════
    # Sheet 2: 生成参数
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet('生成参数')

    headers2 = ['参数', '值']
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    param_font_key = Font(name='Microsoft YaHei', size=10, bold=True)
    param_font_val = Font(name='Consolas', size=10)

    if params:
        row = 2
        for key, val in params.items():
            cell_k = ws2.cell(row=row, column=1, value=str(key))
            cell_k.font = param_font_key
            cell_k.alignment = Alignment(horizontal='left', vertical='center')
            cell_k.border = thin_border

            cell_v = ws2.cell(row=row, column=2, value=str(val))
            cell_v.font = param_font_val
            cell_v.alignment = Alignment(horizontal='right', vertical='center')
            cell_v.border = thin_border

            row += 1

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 28

    # 保存
    wb.save(filepath)
    print(f"Excel 已导出 → {filepath}", file=sys.stderr)


# ══════════════════════════════════════════════════
# 7. 可视化（原点右下角，X 向左为正，Y 向上为正）
# ══════════════════════════════════════════════════

def visualize(map_w, map_h, ctrl_pts, x_traj, y_traj, output_path,
              n_output=2000, smooth_strength=0.0):
    """
    绘制轨迹可视化图并保存为 PNG。
    坐标系: 原点位于右下角，X 轴向左为正，Y 轴向上为正。
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    # ── 配色（参考 dataviz palette） ──
    SURFACE    = '#fcfcfb'
    TEXT_MAIN  = '#0b0b0b'
    TEXT_SUB   = '#52514e'
    LINE_CLR   = '#2a78d6'
    DOT_CLR    = '#d6762a'
    GRID_CLR   = '#e8e6e1'
    BORDER_CLR = '#bfbdb7'

    # ── 中文字体检测 ──
    _cn_candidates = ['Microsoft YaHei', 'SimHei', 'SimSun',
                      'WenQuanYi Micro Hei', 'Noto Sans CJK SC']
    _has_cjk = False
    try:
        from matplotlib.font_manager import findfont
        _found = findfont(_cn_candidates, fallback_to_default=False)
        # findfont 返回空字符串表示未找到
        if _found and _found != '':
            _has_cjk = True
    except Exception:
        pass

    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        plt.rcParams.update({
            'font.family': 'sans-serif',
            'font.sans-serif': (_cn_candidates if _has_cjk else []) + ['DejaVu Sans'],
            'font.size': 10,
            'axes.edgecolor': BORDER_CLR,
            'axes.facecolor': SURFACE,
            'figure.facecolor': '#ffffff',
            'grid.color': GRID_CLR,
            'grid.linewidth': 0.5,
        })

    # ── 标签文本（无 CJK 字体时回退英文） ──
    smooth_tag = f' | 平滑={smooth_strength:.2f}' if smooth_strength > 0 else ''
    if _has_cjk:
        title_text = (f'丝绸之路轨迹  |  地图 {map_w:.0f} × {map_h:.0f}  |  '
                      f'{len(ctrl_pts)} 个控制点 → {n_output} 等弧长采样点{smooth_tag}')
        x_label = 'X  ←  (原点: 右下角)'
        y_label = 'Y  ↑'
        legend_traj = f'样条轨迹 ({n_output}点)'
        legend_ctrl = f'控制点 ({len(ctrl_pts)}个)'
        legend_start = '起点'
        legend_end = '终点'
    else:
        title_text = (f'Silk Road Trajectory  |  Map {map_w:.0f} × {map_h:.0f}  |  '
                      f'{len(ctrl_pts)} control pts → {n_output} arc-length samples{smooth_tag}')
        x_label = 'X  ←  (origin: bottom-right)'
        y_label = 'Y  ↑'
        legend_traj = f'Spline trajectory ({n_output} pts)'
        legend_ctrl = f'Control points ({len(ctrl_pts)})'
        legend_start = 'Start'
        legend_end = 'End'

    fig, ax = plt.subplots(figsize=(12, 6))

    # ── 坐标系：原点右下角，X 左正，Y 上正 ──
    ax.set_xlim(map_w, 0)
    ax.set_ylim(0, map_h)
    ax.set_aspect('equal')

    # ── 地图边界 ──
    from matplotlib.patches import Rectangle
    border = Rectangle((0, 0), map_w, map_h,
                       fill=False, edgecolor=BORDER_CLR, linewidth=1.2,
                       zorder=0)
    ax.add_patch(border)

    # ── 轨迹线 ──
    ax.plot(x_traj, y_traj, color=LINE_CLR, linewidth=1.6, zorder=2)

    # ── 控制点 ──
    ax.scatter(ctrl_pts[:, 0], ctrl_pts[:, 1],
               color=DOT_CLR, s=28, zorder=3, edgecolors='white',
               linewidths=0.8)

    # ── 起点 / 终点标记 ──
    ax.scatter([x_traj[0]], [y_traj[0]], color='#1a7a1a', s=40,
               zorder=4, marker='s', edgecolors='white', linewidths=0.8)
    ax.scatter([x_traj[-1]], [y_traj[-1]], color='#c42e2e', s=40,
               zorder=4, marker='D', edgecolors='white', linewidths=0.8)

    # ── 网格与刻度 ──
    ax.grid(True, linestyle='-', linewidth=0.5, alpha=0.8)
    ax.tick_params(colors=TEXT_SUB, labelsize=8)
    ax.xaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f'{v:.0f}')
    )
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda v, _: f'{v:.0f}')
    )

    # ── 轴标签与标题 ──
    ax.set_xlabel(x_label, color=TEXT_SUB, fontsize=10)
    ax.set_ylabel(y_label, color=TEXT_SUB, fontsize=10)
    ax.set_title(title_text, color=TEXT_MAIN, fontsize=13, pad=12)

    # ── 图例 ──
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], color=LINE_CLR, linewidth=1.6, label=legend_traj),
        Line2D([0], [0], marker='o', color='w', markerfacecolor=DOT_CLR,
               markersize=7, label=legend_ctrl),
        Line2D([0], [0], marker='s', color='w', markerfacecolor='#1a7a1a',
               markersize=8, label=legend_start),
        Line2D([0], [0], marker='D', color='w', markerfacecolor='#c42e2e',
               markersize=8, label=legend_end),
    ]
    ax.legend(handles=legend_elements, loc='lower left',
              framealpha=0.9, edgecolor=BORDER_CLR, fontsize=8)

    fig.tight_layout()
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', UserWarning)
        fig.savefig(output_path, dpi=150, facecolor='white', edgecolor='none')
    plt.close(fig)
    print(f"可视化已保存 → {output_path}", file=sys.stderr)


# ══════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════

def main():
    if len(sys.argv) >= 2 and sys.argv[1] in ('-h', '--help'):
        print(__doc__)
        sys.exit(0)

    # 解析参数: [--plot] [--smooth S] [--n-output N] [输入文件] [输出文件]
    args = sys.argv[1:]
    do_plot = False
    smooth_strength = 0.0
    n_output = 2000

    # 提取 --smooth 标志
    if '--smooth' in args:
        idx = args.index('--smooth')
        if idx + 1 >= len(args):
            print("错误：--smooth 需要一个参数值", file=sys.stderr)
            sys.exit(1)
        smooth_strength = float(args[idx + 1])
        if smooth_strength < 0 or smooth_strength > 1:
            print("错误：--smooth 参数必须在 [0, 1] 范围内", file=sys.stderr)
            sys.exit(1)
        args.pop(idx)
        args.pop(idx)  # 两次 pop 同一位置：先移除标志，再移除值

    # 提取 --n-output 标志
    if '--n-output' in args:
        idx = args.index('--n-output')
        if idx + 1 >= len(args):
            print("错误：--n-output 需要一个参数值", file=sys.stderr)
            sys.exit(1)
        n_output = int(args[idx + 1])
        if n_output < 10 or n_output > 50000:
            print("错误：--n-output 参数必须在 [10, 50000] 范围内", file=sys.stderr)
            sys.exit(1)
        args.pop(idx)
        args.pop(idx)

    # 提取 --target-length 标志
    target_length = None
    if '--target-length' in args:
        idx = args.index('--target-length')
        if idx + 1 >= len(args):
            print("错误：--target-length 需要一个参数值", file=sys.stderr)
            sys.exit(1)
        target_length = float(args[idx + 1])
        if target_length <= 0:
            print("错误：--target-length 参数必须 > 0", file=sys.stderr)
            sys.exit(1)
        args.pop(idx)
        args.pop(idx)

    if '--plot' in args:
        do_plot = True
        args.remove('--plot')

    do_xlsx = False
    if '--xlsx' in args:
        do_xlsx = True
        args.remove('--xlsx')

    input_file = args[0] if len(args) >= 1 else None
    output_file = args[1] if len(args) >= 2 else None

    # Step 1: 读取地图尺寸与控制点
    if input_file is None:
        print("请输入地图尺寸（宽度 高度），然后输入控制点数量 N，"
              "再逐行输入 N 个控制点坐标（X Y），"
              "输入 Ctrl+Z 后回车结束:",
              file=sys.stderr)
    (map_w, map_h), raw_pts = read_map_and_points(input_file)
    n_pts = len(raw_pts)
    print(f"地图尺寸: {map_w:.0f} × {map_h:.0f}", file=sys.stderr)
    print(f"控制点数量: {n_pts}", file=sys.stderr)

    # Step 2: 标准弦长参数化（始终使用原始控制点，不做预平滑）
    if smooth_strength > 0:
        print(f"平滑强度: {smooth_strength:.2f}（后处理模式：轨迹始终经过所有控制点）",
              file=sys.stderr)

    chords = np.sqrt(np.sum(np.diff(raw_pts, axis=0) ** 2, axis=1))
    chord_sum = float(np.sum(chords))
    t_knots = np.zeros(n_pts, dtype=np.float64)
    t_knots[1:] = np.cumsum(chords)

    if t_knots[-1] < 1e-12:
        print("错误：控制点不能全部相同", file=sys.stderr)
        sys.exit(1)

    x_ctrl = raw_pts[:, 0]
    y_ctrl = raw_pts[:, 1]

    # Step 3: 拟合自然三次样条（始终经过原始控制点）
    ax_spline, bx_s, cx_s, dx_s = fit_natural_cubic_spline(t_knots, x_ctrl)
    ay_spline, by_s, cy_s, dy_s = fit_natural_cubic_spline(t_knots, y_ctrl)

    # Step 4: 若有目标长度，二分搜索 α 并混合系数
    alpha_opt = 0.0
    if target_length is not None:
        print(f"期望轨迹长度: {target_length:.2f}  (弦长总和: {chord_sum:.2f})",
              file=sys.stderr)
        alpha_opt, L_achieved, success, warn_msg = find_alpha_for_target_length(
            t_knots, ax_spline, bx_s, cx_s, dx_s,
            ay_spline, by_s, cy_s, dy_s,
            x_ctrl, y_ctrl,
            target_length, chord_sum
        )
        if warn_msg:
            print(warn_msg, file=sys.stderr)
        if success:
            print(f"  最优 α = {alpha_opt:.4f}  →  弧长 = {L_achieved:.2f} "
                  f"(误差 {abs(L_achieved - target_length):.2f})", file=sys.stderr)
        else:
            print(f"  使用 α = {alpha_opt:.4f}  →  弧长 = {L_achieved:.2f}",
                  file=sys.stderr)

    # 混合样条系数（α=0 即保持原样条不变）
    if abs(alpha_opt) > 1e-10:
        ax_spline, bx_s, cx_s, dx_s = blend_spline_coeffs(
            t_knots, ax_spline, bx_s, cx_s, dx_s, x_ctrl, alpha_opt)
        ay_spline, by_s, cy_s, dy_s = blend_spline_coeffs(
            t_knots, ay_spline, by_s, cy_s, dy_s, y_ctrl, alpha_opt)

    # Step 5: 等弧长重采样
    print(f"输出点数: {n_output}", file=sys.stderr)
    x_out, y_out, xp, yp, xpp, ypp = arc_length_reparam(
        t_knots, ax_spline, bx_s, cx_s, dx_s, ay_spline, by_s, cy_s, dy_s,
        n_output=n_output, n_dense=20000
    )

    # Step 5.5: 后处理平滑 + 控制点对齐（控制点位置始终不变）
    if smooth_strength > 0:
        x_out, y_out = smooth_output_trajectory(
            x_out, y_out, raw_pts, smooth_strength)
        # 平滑后使用数值导数
        xp = np.gradient(x_out)
        yp = np.gradient(y_out)
        xpp = np.gradient(xp)
        ypp = np.gradient(yp)
    else:
        # 即使不平滑，也将最近输出点对齐到控制点（弥补离散采样误差）
        for cp in raw_pts:
            dist2 = (x_out - cp[0]) ** 2 + (y_out - cp[1]) ** 2
            idx = int(np.argmin(dist2))
            x_out[idx] = cp[0]
            y_out[idx] = cp[1]

    # Step 6: 计算有符号曲率半径
    R_out = compute_signed_radius(xp, yp, xpp, ypp)

    # Step 7: 输出（确定数据文件与可视化文件路径）
    data_file = output_file
    plot_path = None

    if do_plot:
        import os
        if output_file:
            if output_file.endswith('.png'):
                # 用户指定 .png 为主输出 → 数据写同名 .txt
                data_file = os.path.splitext(output_file)[0] + '.txt'
                plot_path = output_file
            else:
                plot_path = os.path.splitext(output_file)[0] + '.png'
        else:
            plot_path = 'trajectory.png'

    output_results(x_out, y_out, R_out, data_file)

    if data_file:
        print(f"已输出 {n_output} 行数据 → {data_file}", file=sys.stderr)
    else:
        print(f"已输出 {n_output} 行数据（stdout）", file=sys.stderr)

    # Step 7.5: 导出 Excel（可选）
    if do_xlsx:
        import os
        # 计算总弧长
        arc_len = float(np.sum(np.sqrt(
            np.diff(x_out) ** 2 + np.diff(y_out) ** 2
        )))
        # 确定 xlsx 路径
        if output_file and not output_file.endswith('.png'):
            xlsx_path = os.path.splitext(output_file)[0] + '.xlsx'
        elif data_file:
            xlsx_path = os.path.splitext(data_file)[0] + '.xlsx'
        else:
            xlsx_path = 'guiji_XYR.xlsx'

        params = {
            '地图宽度': f'{map_w:.0f}',
            '地图高度': f'{map_h:.0f}',
            '控制点数量': str(n_pts),
            '输出点数': str(n_output),
            '平滑强度': f'{smooth_strength:.2f}',
            '弦长总和': f'{chord_sum:.2f}',
            '轨迹总弧长': f'{arc_len:.2f}',
            '系数混合 α': f'{alpha_opt:.4f}',
        }
        if target_length is not None:
            params['期望长度'] = f'{target_length:.2f}'

        output_xlsx(xlsx_path, x_out, y_out, R_out, raw_pts, params)

    # Step 8: 可视化（可选）
    if do_plot:
        visualize(map_w, map_h, raw_pts, x_out, y_out, plot_path,
                  n_output=n_output, smooth_strength=smooth_strength)


if __name__ == "__main__":
    main()
